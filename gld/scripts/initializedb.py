from __future__ import unicode_literals

import os
import sys

from openpyxl import load_workbook
from clld.scripts.util import initializedb, Data
from clld.db.models import common
from clld.db.meta import DBSession

from gld.models import Word


def main(args):
    data = Data()
    data.add(
        common.Dataset,
        "starling",
        id="starling",
        name="Starling",
        domain="starling.rinet.ru",
        description=" The Global Lexicostatistical Database",
        publisher_name="Russian State University for the Humanities, Moscow")
    data.add(common.Contribution, "starling", name="Starling", id="starling")

    def row_to_dict(row_entry):
        swadesh_id_idx, swadesh_word_idx, form_idx, cognation_idx, notes_idx = range(0, 5)
        return {
            "swadesh_id": row_entry[swadesh_id_idx].value,
            "swadesh_word": row_entry[swadesh_word_idx].value,
            "form": row_entry[form_idx].value,
            "cognation_index": row_entry[cognation_idx].value,
            "notes": row_entry[notes_idx].value,
        }

    data_dir = "./gld/scripts/data/"
    for path in os.listdir(data_dir):
        data_file_path = os.path.join(data_dir, path)
        book = load_workbook(data_file_path)
        sheet = book.active
        lang_name = sheet["C1"].value
        data.add(common.Language, lang_name, id=lang_name, name=lang_name, latitude=52.0, longitude=0.0)

        fields = [
            "swadesh_id",
            "swadesh_word",
            "form",
            "cognation_index",
            "notes"
        ]
        for row in sheet.iter_rows(min_row=2, min_col=1):
            row_data = row_to_dict(row)
            w = data.add(
                Word,
                "%s_%s" % (row_data["swadesh_id"], row_data["form"]),
                name=row_data["form"],
                description="Description",
                jsondata={k: row_data[k] for k in fields})
            w.language = data["Language"][lang_name]

        DBSession.flush()


def prime_cache(args):
    """If data needs to be denormalized for lookup, do that here.
    This procedure should be separate from the db initialization, because
    it will have to be run periodiucally whenever data has been updated.
    """


if __name__ == "__main__":  # pragma: no cover
    initializedb(create=main, prime_cache=prime_cache)
    sys.exit(0)
